"""
Script de Scraping - MÓDULO QUALIDADE DO MOBUSS
Extrai todos os recursos da seção Qualidade
"""

import json
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Configurações
URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/portal/home/iu00_02.jsf"
EMAIL = "ti@grupoinvestcorp.com.br"
SENHA = "FT$0luc0#$TI"

class MobussQualidadeScraper:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {
            "timestamp": datetime.now().isoformat(),
            "secoes": {},
            "formularios": [],
            "campos": [],
            "opcoes": [],
            "tabelas": [],
            "estrutura_menu": []
        }

    def setup_driver(self):
        """Configura o driver do Selenium"""
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, 10)
        print("✅ Driver iniciado")

    def login(self):
        """Realiza login no Mobuss"""
        try:
            print(f"🔐 Acessando: {URL}")
            self.driver.get(URL)
            time.sleep(3)

            # Tenta encontrar campos de login
            email_field = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='email'], input[type='text'][name*='usuario']"))
            )
            senha_field = self.driver.find_element(By.CSS_SELECTOR, "input[type='password']")

            email_field.clear()
            email_field.send_keys(EMAIL)
            time.sleep(1)

            senha_field.clear()
            senha_field.send_keys(SENHA)
            time.sleep(1)

            # Clica em login
            login_btn = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")
            login_btn.click()

            time.sleep(5)
            print("✅ Login realizado")

        except Exception as e:
            print(f"❌ Erro no login: {e}")
            self.captura_screenshot("erro_login.png")
            raise

    def acessar_qualidade(self):
        """Acessa a seção Qualidade no menu"""
        try:
            print("\n📍 Acessando seção QUALIDADE...")

            # Clica no menu Qualidade
            qualidade_menu = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Qualidade')]"))
            )
            qualidade_menu.click()
            time.sleep(2)

            print("✅ Seção Qualidade acessada")

        except Exception as e:
            print(f"❌ Erro ao acessar Qualidade: {e}")

    def extrair_secoes_menu(self):
        """Extrai todas as subseções do menu Qualidade"""
        try:
            print("\n📋 Extraindo seções do menu...")

            # Procura todos os items de menu em Qualidade
            menu_items = self.driver.find_elements(By.XPATH, "//nav//a, //menu//a, //div[@class*='menu']//a")

            secoes = []
            for item in menu_items:
                texto = item.text.strip()
                href = item.get_attribute("href")

                if texto and href:
                    secoes.append({
                        "nome": texto,
                        "url": href
                    })

            self.data["estrutura_menu"] = secoes
            print(f"✅ {len(secoes)} seção(ões) encontrada(s)")

            for secao in secoes:
                print(f"  - {secao['nome']}")

        except Exception as e:
            print(f"⚠️ Erro ao extrair seções: {e}")

    def navegar_e_extrair_secoes(self):
        """Navega por cada seção e extrai dados"""
        secoes_importantes = [
            "Painel de Qualidade",
            "Cobertura",
            "Modelos de Análise",
            "Acompanhamento de Inspeções",
            "Realizados",
            "Cadastros",
            "Tipos de Formulários",
            "Gestão",
            "Ações",
            "Tipos"
        ]

        for secao in self.data["estrutura_menu"]:
            nome_secao = secao["nome"]

            # Processa apenas seções importantes
            if any(importante.lower() in nome_secao.lower() for importante in secoes_importantes):
                try:
                    print(f"\n🔄 Processando: {nome_secao}")

                    # Navega para a seção
                    self.driver.get(secao["url"])
                    time.sleep(2)

                    # Extrai dados da seção
                    dados_secao = self.extrair_dados_secao(nome_secao)
                    self.data["secoes"][nome_secao] = dados_secao

                    # Captura screenshot
                    self.captura_screenshot(f"secao_{nome_secao.replace(' ', '_').lower()}.png")

                    print(f"✅ Seção processada")

                except Exception as e:
                    print(f"⚠️ Erro ao processar {nome_secao}: {e}")

    def extrair_dados_secao(self, nome_secao):
        """Extrai todos os dados de uma seção específica"""
        dados = {
            "url": self.driver.current_url,
            "titulo": self.driver.title,
            "formularios": [],
            "tabelas": [],
            "campos": [],
            "opcoes": []
        }

        try:
            # Extrai formulários
            formularios = self.driver.find_elements(By.TAG_NAME, "form")
            for form in formularios:
                form_data = {
                    "id": form.get_attribute("id"),
                    "campos": self.extrair_campos_form(form)
                }
                dados["formularios"].append(form_data)

            # Extrai tabelas
            tabelas = self.driver.find_elements(By.TAG_NAME, "table")
            for idx, tabela in enumerate(tabelas):
                tabela_data = {
                    "numero": idx + 1,
                    "headers": [],
                    "linhas": [],
                    "total_linhas": 0
                }

                # Headers
                headers = tabela.find_elements(By.TAG_NAME, "th")
                tabela_data["headers"] = [h.text for h in headers]

                # Linhas
                rows = tabela.find_elements(By.TAG_NAME, "tr")
                for row in rows[1:]:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    linha = [cell.text for cell in cells]
                    if linha:
                        tabela_data["linhas"].append(linha)

                tabela_data["total_linhas"] = len(tabela_data["linhas"])
                dados["tabelas"].append(tabela_data)

            # Extrai selects e opções
            selects = self.driver.find_elements(By.TAG_NAME, "select")
            for select in selects:
                options = select.find_elements(By.TAG_NAME, "option")
                opcoes_list = [{"value": opt.get_attribute("value"), "text": opt.text} for opt in options]

                dados["opcoes"].append({
                    "select_name": select.get_attribute("name"),
                    "opcoes": opcoes_list
                })

        except Exception as e:
            print(f"⚠️ Erro ao extrair dados da seção: {e}")

        return dados

    def extrair_campos_form(self, form):
        """Extrai campos de um formulário"""
        campos = []

        try:
            # Inputs
            inputs = form.find_elements(By.TAG_NAME, "input")
            for inp in inputs:
                campos.append({
                    "tipo": "input",
                    "input_type": inp.get_attribute("type"),
                    "name": inp.get_attribute("name"),
                    "id": inp.get_attribute("id"),
                    "placeholder": inp.get_attribute("placeholder")
                })

            # Selects
            selects = form.find_elements(By.TAG_NAME, "select")
            for select in selects:
                opcoes = [{"value": opt.get_attribute("value"), "text": opt.text}
                         for opt in select.find_elements(By.TAG_NAME, "option")]

                campos.append({
                    "tipo": "select",
                    "name": select.get_attribute("name"),
                    "opcoes": opcoes
                })

            # Textareas
            textareas = form.find_elements(By.TAG_NAME, "textarea")
            for textarea in textareas:
                campos.append({
                    "tipo": "textarea",
                    "name": textarea.get_attribute("name"),
                    "id": textarea.get_attribute("id")
                })

        except Exception as e:
            print(f"⚠️ Erro ao extrair campos: {e}")

        return campos

    def captura_screenshot(self, nome):
        """Captura screenshot da página"""
        try:
            caminho = os.path.join("screenshots", nome)
            os.makedirs("screenshots", exist_ok=True)
            self.driver.save_screenshot(caminho)
            print(f"  📸 Screenshot: {nome}")
        except Exception as e:
            print(f"⚠️ Erro ao capturar screenshot: {e}")

    def exportar_json(self):
        """Exporta dados em JSON"""
        try:
            caminho = "relatorio_mobuss_qualidade.json"
            with open(caminho, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            print(f"\n✅ JSON exportado: {caminho}")
        except Exception as e:
            print(f"❌ Erro ao exportar JSON: {e}")

    def exportar_excel(self):
        """Exporta dados em Excel"""
        try:
            wb = Workbook()

            # Aba 1: Resumo de Seções
            ws1 = wb.active
            ws1.title = "Resumo"
            ws1['A1'] = "Seção"
            ws1['B1'] = "URL"
            ws1['C1'] = "Formulários"
            ws1['D1'] = "Tabelas"
            ws1['E1'] = "Campos"
            ws1['F1'] = "Opções"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                ws1[f'A{row}'] = secao_nome
                ws1[f'B{row}'] = dados_secao.get("url", "")
                ws1[f'C{row}'] = len(dados_secao.get("formularios", []))
                ws1[f'D{row}'] = len(dados_secao.get("tabelas", []))
                ws1[f'E{row}'] = len(dados_secao.get("campos", []))
                ws1[f'F{row}'] = len(dados_secao.get("opcoes", []))
                row += 1

            # Aba 2: Campos Encontrados
            ws2 = wb.create_sheet("Campos")
            ws2['A1'] = "Seção"
            ws2['B1'] = "Tipo"
            ws2['C1'] = "Nome"
            ws2['D1'] = "ID"
            ws2['E1'] = "Placeholder"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                for campo in dados_secao.get("campos", []):
                    ws2[f'A{row}'] = secao_nome
                    ws2[f'B{row}'] = campo.get("tipo", "")
                    ws2[f'C{row}'] = campo.get("name", "")
                    ws2[f'D{row}'] = campo.get("id", "")
                    ws2[f'E{row}'] = campo.get("placeholder", "")
                    row += 1

            # Aba 3: Opções de Dropdowns
            ws3 = wb.create_sheet("Opções")
            ws3['A1'] = "Seção"
            ws3['B1'] = "Select"
            ws3['C1'] = "Opção"
            ws3['D1'] = "Value"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                for select_data in dados_secao.get("opcoes", []):
                    for opcao in select_data.get("opcoes", []):
                        ws3[f'A{row}'] = secao_nome
                        ws3[f'B{row}'] = select_data.get("select_name", "")
                        ws3[f'C{row}'] = opcao.get("text", "")
                        ws3[f'D{row}'] = opcao.get("value", "")
                        row += 1

            # Aba 4: Tabelas
            ws4 = wb.create_sheet("Tabelas")
            ws4['A1'] = "Seção"
            ws4['B1'] = "Tabela"
            ws4['C1'] = "Headers"
            ws4['D1'] = "Linhas"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                for tabela in dados_secao.get("tabelas", []):
                    ws4[f'A{row}'] = secao_nome
                    ws4[f'B{row}'] = tabela.get("numero", "")
                    ws4[f'C{row}'] = ", ".join(tabela.get("headers", []))
                    ws4[f'D{row}'] = tabela.get("total_linhas", 0)
                    row += 1

            caminho = "relatorio_mobuss_qualidade.xlsx"
            wb.save(caminho)
            print(f"✅ Excel exportado: {caminho}")

        except Exception as e:
            print(f"❌ Erro ao exportar Excel: {e}")

    def executar(self):
        """Executa o scraping completo"""
        try:
            print("🚀 Iniciando scraping do MÓDULO QUALIDADE do Mobuss...\n")

            self.setup_driver()
            self.login()
            self.acessar_qualidade()
            self.extrair_secoes_menu()
            self.navegar_e_extrair_secoes()

            print("\n💾 Exportando relatórios...\n")
            self.exportar_json()
            self.exportar_excel()

            print("\n✅ Scraping concluído com sucesso!")
            print("\n📊 Arquivos gerados:")
            print("  - relatorio_mobuss_qualidade.json")
            print("  - relatorio_mobuss_qualidade.xlsx")
            print("  - screenshots/ (múltiplas imagens)")

        except Exception as e:
            print(f"\n❌ Erro geral: {e}")

        finally:
            if self.driver:
                self.driver.quit()
                print("\n🔌 Driver fechado")

if __name__ == "__main__":
    scraper = MobussQualidadeScraper()
    scraper.executar()
