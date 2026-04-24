#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scraper Mobuss - Com retry e delays maiores para Keycloak
"""

import json
import time
import sys
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import os

LOGIN_URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/portal/home/iu00_02.jsf"
QUALIDADE_URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/qualidade/checklist/realizado/iu01_23.jsf"

EMAIL = "ti@grupoinvestcorp.com.br"
SENHA = "FT$0luc0#$TI"

class MobussScraper:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {
            "timestamp": datetime.now().isoformat(),
            "secoes": {}
        }

    def log(self, nivel, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] [{nivel}] {msg}")
        sys.stdout.flush()

    def setup_driver(self):
        try:
            opts = Options()
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            self.driver = webdriver.Chrome(options=opts)
            self.wait = WebDriverWait(self.driver, 20)
            self.log("OK", "Driver iniciado")
        except Exception as e:
            self.log("ERRO", f"Setup: {e}")
            raise

    def login(self):
        """Login com retry e delays aumentados"""
        try:
            self.log("INFO", "Acessando login...")
            self.driver.get(LOGIN_URL)
            time.sleep(4)

            # ETAPA 1: Username
            self.log("INFO", "Etapa 1: Preenchendo usuario...")
            user_input = self.wait.until(
                EC.presence_of_element_located((By.NAME, "username"))
            )
            user_input.clear()
            user_input.send_keys(EMAIL)
            time.sleep(2)

            # Clica em Entrar
            self.log("INFO", "Etapa 1: Clicando em Entrar...")
            btn = self.driver.find_element(By.NAME, "login")
            btn.click()

            # AGUARDA PROCESSAMENTO - Aguarda até 15 segundos
            self.log("INFO", "Etapa 1: Aguardando processamento do servidor (até 15s)...")
            start_time = time.time()
            while time.time() - start_time < 15:
                time.sleep(1)

                # Verifica se apareceu campo de senha
                try:
                    pwd_field = self.driver.find_element(By.NAME, "password")
                    self.log("OK", "Campo de senha apareceu!")
                    time.sleep(2)
                    break
                except:
                    # Campo ainda não existe, continua aguardando
                    elapsed = int(time.time() - start_time)
                    print(f"[{datetime.now().strftime('%H:%M:%S')}] [INFO]   Aguardando... ({elapsed}s)")

            # ETAPA 2: Password
            self.log("INFO", "Etapa 2: Procurando campo de senha...")
            try:
                pwd_input = self.driver.find_element(By.NAME, "password")
                self.log("OK", "Campo de senha encontrado!")

                self.log("INFO", "Etapa 2: Preenchendo senha...")
                pwd_input.clear()
                pwd_input.send_keys(SENHA)
                time.sleep(2)

                # Clica em Entrar novamente
                self.log("INFO", "Etapa 2: Clicando em Entrar...")
                btn2 = self.driver.find_element(By.NAME, "login")
                btn2.click()

                # AGUARDA REDIRECIONAMENTO - até 15 segundos
                self.log("INFO", "Etapa 2: Aguardando autenticacao (até 15s)...")
                start_time = time.time()
                while time.time() - start_time < 15:
                    time.sleep(1)

                    url_atual = self.driver.current_url
                    if "login" not in url_atual.lower():
                        self.log("OK", "Autenticacao bem-sucedida!")
                        time.sleep(2)
                        break
                    else:
                        elapsed = int(time.time() - start_time)
                        print(f"[{datetime.now().strftime('%H:%M:%S')}] [INFO]   Processando... ({elapsed}s)")

            except Exception as e:
                self.log("AVISO", f"Sem campo de senha ou erro: {str(e)[:80]}")

            # Verifica resultado final
            url_final = self.driver.current_url
            titulo_final = self.driver.title
            self.log("INFO", f"URL final: {url_final[:80]}")
            self.log("INFO", f"Titulo: {titulo_final[:80]}")

            if "login" in url_final.lower():
                self.log("AVISO", "Ainda na pagina de login - autenticacao pode ter falhado")
                self.screenshot("erro_login_ainda.png")
            else:
                self.log("OK", "Login aparentemente bem-sucedido!")

        except Exception as e:
            self.log("ERRO", f"Login: {str(e)[:150]}")
            self.screenshot("erro_login_geral.png")
            raise

    def acessar_qualidade(self):
        try:
            self.log("INFO", "Acessando Qualidade...")
            url_atual = self.driver.current_url

            if "qualidade" not in url_atual.lower():
                self.log("INFO", "Navegando para URL de Qualidade...")
                self.driver.get(QUALIDADE_URL)

            time.sleep(5)

            url = self.driver.current_url
            titulo = self.driver.title

            self.log("INFO", f"URL: {url[:80]}")
            self.log("INFO", f"Titulo: {titulo[:80]}")

            if "qualidade" in url.lower() and "login" not in url.lower():
                self.log("OK", "Qualidade acessada com sucesso!")
            else:
                self.log("AVISO", "Pode estar em pagina incorreta")

        except Exception as e:
            self.log("ERRO", f"Qualidade: {e}")

    def extrair_dados(self):
        try:
            self.log("INFO", "Extraindo dados...")

            titulo = self.driver.title
            url = self.driver.current_url

            # Formularios
            forms = self.driver.find_elements(By.TAG_NAME, "form")
            form_count = len(forms)

            # Inputs
            all_inputs = self.driver.find_elements(By.TAG_NAME, "input")
            input_data = []
            for inp in all_inputs:
                itype = inp.get_attribute("type")
                name = inp.get_attribute("name")
                if itype not in ["submit", "button", "hidden"] and name:
                    input_data.append({
                        "type": itype,
                        "name": name,
                        "id": inp.get_attribute("id"),
                        "placeholder": inp.get_attribute("placeholder")
                    })

            # Selects
            selects = self.driver.find_elements(By.TAG_NAME, "select")
            select_data = []
            for sel in selects:
                opts = sel.find_elements(By.TAG_NAME, "option")
                opt_texts = [opt.text.strip() for opt in opts if opt.text.strip()]
                select_data.append({
                    "name": sel.get_attribute("name"),
                    "id": sel.get_attribute("id"),
                    "options": opt_texts
                })

            # Tabelas
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            table_data = []
            for tbl in tables:
                ths = tbl.find_elements(By.TAG_NAME, "th")
                trs = tbl.find_elements(By.TAG_NAME, "tr")
                headers = [th.text.strip() for th in ths if th.text.strip()]
                table_data.append({
                    "headers": headers,
                    "rows": len(trs) - 1 if trs else 0
                })

            # Textareas
            textareas = self.driver.find_elements(By.TAG_NAME, "textarea")

            self.log("OK", f"Dados extraidos: Forms={form_count}, Inputs={len(input_data)}, Selects={len(select_data)}, Tabelas={len(table_data)}")

            self.data["secoes"]["Qualidade - Realizados"] = {
                "url": url,
                "titulo": titulo,
                "formularios": form_count,
                "inputs": input_data[:50],
                "selects": select_data[:30],
                "tabelas": table_data[:15],
                "textareas": len(textareas)
            }

            self.screenshot("secao_qualidade_final.png")

        except Exception as e:
            self.log("AVISO", f"Extracao: {str(e)[:100]}")

    def screenshot(self, nome):
        try:
            os.makedirs("screenshots", exist_ok=True)
            caminho = os.path.join("screenshots", nome)
            self.driver.save_screenshot(caminho)
            self.log("INFO", f"Screenshot: {nome}")
        except:
            pass

    def exportar_json(self):
        try:
            with open("relatorio_mobuss_qualidade.json", 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            self.log("OK", "JSON exportado")
        except Exception as e:
            self.log("ERRO", f"JSON: {e}")

    def exportar_excel(self):
        try:
            wb = Workbook()

            # Aba 1: Resumo
            ws1 = wb.active
            ws1.title = "Resumo"
            ws1['A1'] = "Parametro"
            ws1['B1'] = "Valor"

            row = 2
            for secao, dados in self.data["secoes"].items():
                ws1[f'A{row}'] = "Secao"
                ws1[f'B{row}'] = secao
                row += 1

                ws1[f'A{row}'] = "URL"
                ws1[f'B{row}'] = dados.get("url", "")
                row += 1

                ws1[f'A{row}'] = "Formularios"
                ws1[f'B{row}'] = dados.get("formularios", 0)
                row += 1

                ws1[f'A{row}'] = "Inputs"
                ws1[f'B{row}'] = len(dados.get("inputs", []))
                row += 1

                ws1[f'A{row}'] = "Selects"
                ws1[f'B{row}'] = len(dados.get("selects", []))
                row += 1

                ws1[f'A{row}'] = "Tabelas"
                ws1[f'B{row}'] = len(dados.get("tabelas", []))
                row += 1

                ws1[f'A{row}'] = "Textareas"
                ws1[f'B{row}'] = dados.get("textareas", 0)
                row += 2

            # Aba 2: Inputs
            ws2 = wb.create_sheet("Inputs")
            ws2['A1'] = "Secao"
            ws2['B1'] = "Tipo"
            ws2['C1'] = "Nome"
            ws2['D1'] = "ID"
            ws2['E1'] = "Placeholder"

            row = 2
            for secao, dados in self.data["secoes"].items():
                for inp in dados.get("inputs", []):
                    ws2[f'A{row}'] = secao
                    ws2[f'B{row}'] = inp.get("type", "")
                    ws2[f'C{row}'] = inp.get("name", "")
                    ws2[f'D{row}'] = inp.get("id", "")
                    ws2[f'E{row}'] = inp.get("placeholder", "")
                    row += 1

            # Aba 3: Selects
            ws3 = wb.create_sheet("Selects")
            ws3['A1'] = "Secao"
            ws3['B1'] = "Nome"
            ws3['C1'] = "Opcoes (primeiras 5)"

            row = 2
            for secao, dados in self.data["secoes"].items():
                for sel in dados.get("selects", []):
                    ws3[f'A{row}'] = secao
                    ws3[f'B{row}'] = sel.get("name", "")
                    ws3[f'C{row}'] = "; ".join(sel.get("options", [])[:5])
                    row += 1

            # Aba 4: Tabelas
            ws4 = wb.create_sheet("Tabelas")
            ws4['A1'] = "Secao"
            ws4['B1'] = "Headers"
            ws4['C1'] = "Total Linhas"

            row = 2
            for secao, dados in self.data["secoes"].items():
                for idx, tbl in enumerate(dados.get("tabelas", []), 1):
                    ws4[f'A{row}'] = f"{secao} - Tabela {idx}"
                    ws4[f'B{row}'] = "; ".join(tbl.get("headers", [])[:10])
                    ws4[f'C{row}'] = tbl.get("rows", 0)
                    row += 1

            wb.save("relatorio_mobuss_qualidade.xlsx")
            self.log("OK", "Excel exportado")

        except Exception as e:
            self.log("ERRO", f"Excel: {e}")

    def executar(self):
        try:
            self.log("INFO", "=== SCRAPER MOBUSS QUALIDADE COM RETRY ===\n")

            self.setup_driver()
            self.login()
            time.sleep(3)
            self.acessar_qualidade()
            time.sleep(3)
            self.extrair_dados()

            self.log("INFO", "\n=== EXPORTANDO ===\n")
            self.exportar_json()
            self.exportar_excel()

            self.log("OK", "Scraping concluido!")

        except Exception as e:
            self.log("ERRO", f"Geral: {e}")
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass

if __name__ == "__main__":
    scraper = MobussScraper()
    scraper.executar()
