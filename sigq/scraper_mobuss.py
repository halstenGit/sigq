"""
Script de Scraping do Mobuss - VERSÃO MOCK COM DADOS FICTÍCIOS
TODO: Integrar com API real quando disponível

Este script atualmente usa dados fictícios para desenvolvimento.
Será substituído por integração com API quando pronta.
"""

import json
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Configurações - TODO: Mover para .env
# URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/qualidade/checklist/realizado/iu01_23.jsf"
# EMAIL = os.getenv("MOBUSS_EMAIL", "user@example.com")
# SENHA = os.getenv("MOBUSS_PASSWORD", "")

class MobussScraper:
    def __init__(self):
        """Inicializa scraper com dados fictícios"""
        self.data = {
            "timestamp": datetime.now().isoformat(),
            "formularios": [],
            "campos": [],
            "opcoes": [],
            "empreendimentos": [],
            "servicos": [],
            "fvs_extraidas": [],
            "rncs_extraidas": [],
            "estrutura": {},
        }
        self._carregar_dados_ficticios()

    def _carregar_dados_ficticios(self):
        """Carrega dados fictícios para desenvolvimento"""
        print("[MOCK] Carregando dados fictícios de FVS...")

        # Empreendimentos fictícios
        self.data["empreendimentos"] = [
            {"id": "E001", "nome": "Soul Residencial", "cidade": "Joinville"},
            {"id": "E002", "nome": "Morada de Gaia", "cidade": "Joinville"},
            {"id": "E003", "nome": "Ocean View", "cidade": "Balneário Camboriú"},
        ]

        # Serviços fictícios
        self.data["servicos"] = [
            {"id": "S001", "nome": "Estrutura"},
            {"id": "S002", "nome": "Alvenaria"},
            {"id": "S003", "nome": "Revestimento Cerâmico"},
            {"id": "S004", "nome": "Pintura"},
            {"id": "S005", "nome": "Impermeabilização"},
        ]

        # FVS fictícias extraídas
        self.data["fvs_extraidas"] = [
            {
                "id": "FVS-MOCK-0001",
                "empreendimento": "Soul Residencial",
                "servico": "Alvenaria",
                "local": "BL A - 3º PAV",
                "nota": 8.5,
                "status": "finalizada",
                "dataRealizacao": "25/04/2026",
                "inspetor": "Marcelo Sena",
            },
            {
                "id": "FVS-MOCK-0002",
                "empreendimento": "Morada de Gaia",
                "servico": "Revestimento Cerâmico",
                "local": "BL B - 2º PAV",
                "nota": 6.0,
                "status": "com_nc",
                "dataRealizacao": "24/04/2026",
                "inspetor": "Amanda Costa",
            },
            {
                "id": "FVS-MOCK-0003",
                "empreendimento": "Ocean View",
                "servico": "Pintura",
                "local": "BL C - Cobertura",
                "nota": 9.2,
                "status": "finalizada",
                "dataRealizacao": "23/04/2026",
                "inspetor": "Francisco Neto",
            },
        ]

        # RNCs fictícias vinculadas a FVS
        self.data["rncs_extraidas"] = [
            {
                "id": "RNC-MOCK-0001",
                "fvsId": "FVS-MOCK-0002",
                "empreendimento": "Morada de Gaia",
                "servico": "Revestimento Cerâmico",
                "descricao": "Juntas irregulares no banheiro, variacao acima do tolerado",
                "gravidade": "maior",
                "status": "em_correcao",
                "responsavel": "Construtora Alfa",
                "abertura": "24/04/2026",
                "prazo": "28/04/2026",
            },
        ]

    def setup_driver(self):
        """Mock - Não precisa de driver"""
        print("[OK] Modo offline ativado (dados fictícios)")

    def login(self):
        """Mock - Login simulado"""
        print("[OK] Autenticação simulada (dados fictícios)")
        time.sleep(1)

    def extrair_formularios(self):
        """Mock - Retorna estrutura de formulário fictício"""
        self.data["formularios"] = [
            {
                "id": "form_fvs",
                "name": "formFVS",
                "action": "/submit_fvs",
                "method": "POST",
                "campos": [
                    {"tipo": "text", "name": "empreendimento", "required": True},
                    {"tipo": "text", "name": "servico", "required": True},
                    {"tipo": "text", "name": "local", "required": True},
                    {"tipo": "number", "name": "nota", "required": True},
                    {"tipo": "textarea", "name": "observacoes"},
                ]
            }
        ]
        print(f"[OK] {len(self.data['formularios'])} formulário(s) carregado(s)")

    def extrair_empreendimentos(self):
        """Extrai empreendimentos disponíveis"""
        print(f"[OK] {len(self.data['empreendimentos'])} empreendimento(s) carregado(s)")
        return self.data["empreendimentos"]

    def extrair_servicos(self):
        """Extrai serviços disponíveis"""
        print(f"[OK] {len(self.data['servicos'])} serviço(s) carregado(s)")
        return self.data["servicos"]

    def extrair_fvs(self):
        """Extrai FVS disponíveis"""
        print(f"[OK] {len(self.data['fvs_extraidas'])} FVS carregada(s)")
        return self.data["fvs_extraidas"]

    def extrair_rncs(self):
        """Extrai RNCs vinculadas a FVS"""
        print(f"[OK] {len(self.data['rncs_extraidas'])} RNC(s) carregada(s)")
        return self.data["rncs_extraidas"]

    def extrair_tabelas(self):
        """Mock - Retorna dados de tabela fictícia"""
        self.data["estrutura"]["tabelas"] = {
            "total": len(self.data["fvs_extraidas"]),
            "colunas": ["ID", "Empreendimento", "Serviço", "Nota", "Status"]
        }
        print(f"[OK] Tabela com {len(self.data['fvs_extraidas'])} linha(s)")

    def extrair_menu_navegacao(self):
        """Mock - Menu de navegação fictício"""
        self.data["estrutura"]["menu"] = [
            {"text": "Dashboard", "href": "/"},
            {"text": "FVS", "href": "/fvs"},
            {"text": "RNCs", "href": "/rncs"},
        ]
        print(f"[OK] Menu de navegação carregado")

    def extrair_estrutura_pagina(self):
        """Mock - Estrutura fictícia"""
        self.data["estrutura"]["pagina"] = {
            "titulo": "SIGQ - Sistema de Gestão da Qualidade",
            "url": "https://sigq.halsten.local/fvs",
            "versao": "0.1.0"
        }
        print(f"[OK] Estrutura da página carregada")

    def exportar_json(self):
        """Exporta dados em JSON"""
        try:
            caminho = "relatorio_mobuss.json"
            with open(caminho, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            print(f"[OK] JSON exportado: {caminho}")
        except Exception as e:
            print(f"[ERROR] Erro ao exportar JSON: {e}")

    def exportar_excel(self):
        """Exporta dados em Excel"""
        try:
            wb = Workbook()

            # Aba 1: Estrutura
            ws1 = wb.active
            ws1.title = "Estrutura"
            ws1['A1'] = "Página"
            ws1['B1'] = "Valor"
            row = 2
            for chave, valor in self.data["estrutura"].get("pagina", {}).items():
                ws1[f'A{row}'] = chave
                ws1[f'B{row}'] = str(valor)
                row += 1

            # Aba 2: Campos
            ws2 = wb.create_sheet("Campos")
            ws2['A1'] = "Tipo"
            ws2['B1'] = "Nome"
            ws2['C1'] = "ID"
            ws2['D1'] = "Placeholder"
            row = 2
            for campo in self.data["campos"]:
                ws2[f'A{row}'] = campo.get("tipo", "")
                ws2[f'B{row}'] = campo.get("name", "")
                ws2[f'C{row}'] = campo.get("id", "")
                ws2[f'D{row}'] = campo.get("placeholder", "")
                row += 1

            # Aba 3: Opções
            ws3 = wb.create_sheet("Opções")
            ws3['A1'] = "Select"
            ws3['B1'] = "Opção"
            ws3['C1'] = "Value"
            row = 2
            for opcao in self.data["opcoes"]:
                ws3[f'A{row}'] = opcao.get("select", "")
                ws3[f'B{row}'] = opcao.get("option", "")
                ws3[f'C{row}'] = opcao.get("value", "")
                row += 1

            # Aba 4: RNCs
            ws4 = wb.create_sheet("RNCs")
            ws4['A1'] = "ID"
            ws4['B1'] = "FVS ID"
            ws4['C1'] = "Empreendimento"
            ws4['D1'] = "Servico"
            ws4['E1'] = "Descricao"
            ws4['F1'] = "Gravidade"
            ws4['G1'] = "Status"
            ws4['H1'] = "Responsavel"
            ws4['I1'] = "Abertura"
            ws4['J1'] = "Prazo"
            row = 2
            for rnc in self.data["rncs_extraidas"]:
                ws4[f'A{row}'] = rnc.get("id", "")
                ws4[f'B{row}'] = rnc.get("fvsId", "")
                ws4[f'C{row}'] = rnc.get("empreendimento", "")
                ws4[f'D{row}'] = rnc.get("servico", "")
                ws4[f'E{row}'] = rnc.get("descricao", "")
                ws4[f'F{row}'] = rnc.get("gravidade", "")
                ws4[f'G{row}'] = rnc.get("status", "")
                ws4[f'H{row}'] = rnc.get("responsavel", "")
                ws4[f'I{row}'] = rnc.get("abertura", "")
                ws4[f'J{row}'] = rnc.get("prazo", "")
                row += 1

            # Aba 5: Resumo
            ws5 = wb.create_sheet("Resumo")
            ws5['A1'] = "Métrica"
            ws5['B1'] = "Quantidade"
            resumo = [
                ("Formulários", len(self.data["formularios"])),
                ("Campos", len(self.data["campos"])),
                ("Opções", len(self.data["opcoes"])),
                ("Empreendimentos", len(self.data["empreendimentos"])),
                ("Servicos", len(self.data["servicos"])),
                ("FVS", len(self.data["fvs_extraidas"])),
                ("RNCs", len(self.data["rncs_extraidas"])),
            ]
            for idx, (metrica, valor) in enumerate(resumo, start=2):
                ws5[f'A{idx}'] = metrica
                ws5[f'B{idx}'] = valor

            # Salva
            caminho = "relatorio_mobuss.xlsx"
            wb.save(caminho)
            print(f"[OK] Excel exportado: {caminho}")

        except Exception as e:
            print(f"[ERROR] Erro ao exportar Excel: {e}")

    def executar(self):
        """Executa o scraping com dados fictícios"""
        try:
            print(" Iniciando extração de dados (MODO MOCK)...\n")

            self.setup_driver()
            self.login()

            print("\n Extraindo dados...\n")
            self.extrair_estrutura_pagina()
            self.extrair_formularios()
            self.extrair_empreendimentos()
            self.extrair_servicos()
            self.extrair_fvs()
            self.extrair_rncs()
            self.extrair_tabelas()
            self.extrair_menu_navegacao()

            print("\n Exportando relatórios...\n")
            self.exportar_json()
            self.exportar_excel()

            print("\n[OK] Extração concluída com sucesso!")
            print(" Dados fictícios carregados")
            print("[WARN]  TODO: Integrar com API real quando disponível")

        except Exception as e:
            print(f"\n[ERROR] Erro geral: {e}")
            raise

if __name__ == "__main__":
    scraper = MobussScraper()
    scraper.executar()
