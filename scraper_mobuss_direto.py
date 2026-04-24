#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scraper Mobuss - Acesso direto a Qualidade apos login
"""

import json
import time
import sys
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import os

LOGIN_URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/portal/home/iu00_02.jsf"
QUALIDADE_URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/qualidade/checklist/realizado/iu01_23.jsf"

EMAIL = "ti@grupoinvestcorp.com.br"
SENHA = "FT$0luc0#$TI"

class MobussScraper:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {
            "timestamp": datetime.now().isoformat(),
            "login_url": LOGIN_URL,
            "qualidade_url": QUALIDADE_URL,
            "secoes": {}
        }

    def log(self, nivel, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] [{nivel}] {msg}")
        sys.stdout.flush()

    def setup_driver(self):
        try:
            opts = Options()
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            self.driver = webdriver.Chrome(options=opts)
            self.wait = WebDriverWait(self.driver, 12)
            self.log("OK", "Driver iniciado")
        except Exception as e:
            self.log("ERRO", f"Setup: {e}")
            raise

    def login(self):
        try:
            self.log("INFO", f"Acessando login: {LOGIN_URL}")
            self.driver.get(LOGIN_URL)
            time.sleep(3)

            # Campo usuario
            self.log("INFO", "Buscando campo usuario...")
            inputs = self.driver.find_elements(By.TAG_NAME, "input")
            user_input = None

            for inp in inputs:
                itype = inp.get_attribute("type")
                if itype == "text":
                    user_input = inp
                    break

            if not user_input:
                user_input = inputs[0] if inputs else None

            if user_input:
                self.log("INFO", "Preenchendo usuario...")
                user_input.clear()
                user_input.send_keys(EMAIL)
                time.sleep(1)

                # Botao de envio
                btns = self.driver.find_elements(By.TAG_NAME, "button")
                if btns:
                    btns[0].click()
                else:
                    subs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='submit']")
                    if subs:
                        subs[0].click()

                time.sleep(5)
                self.log("OK", "Login etapa 1 concluido")
            else:
                self.log("AVISO", "Campo usuario nao encontrado")

        except Exception as e:
            self.log("ERRO", f"Login: {str(e)[:100]}")
            self.screenshot("erro_login.png")
            raise

    def acessar_qualidade(self):
        try:
            self.log("INFO", f"Acessando qualidade: {QUALIDADE_URL}")
            self.driver.get(QUALIDADE_URL)
            time.sleep(4)
            self.log("OK", "Qualidade acessada")
        except Exception as e:
            self.log("ERRO", f"Qualidade: {e}")
            raise

    def extrair_dados(self):
        try:
            self.log("INFO", "Extraindo dados...")

            # Titulo
            titulo = self.driver.title
            url_atual = self.driver.current_url
            self.log("INFO", f"Titulo: {titulo}")

            # Formularios
            forms = self.driver.find_elements(By.TAG_NAME, "form")
            form_count = len(forms)
            self.log("INFO", f"Formularios: {form_count}")

            # Inputs
            inputs = self.driver.find_elements(By.TAG_NAME, "input")
            input_data = []
            for inp in inputs:
                itype = inp.get_attribute("type")
                name = inp.get_attribute("name")
                if itype not in ["submit", "button"] and name:
                    input_data.append({
                        "type": itype,
                        "name": name,
                        "id": inp.get_attribute("id")
                    })

            self.log("INFO", f"Campos input: {len(input_data)}")

            # Selects
            selects = self.driver.find_elements(By.TAG_NAME, "select")
            select_data = []
            for sel in selects:
                opts = sel.find_elements(By.TAG_NAME, "option")
                opt_texts = [opt.text.strip() for opt in opts]
                select_data.append({
                    "name": sel.get_attribute("name"),
                    "options": opt_texts
                })

            self.log("INFO", f"Selects: {len(select_data)}")

            # Tabelas
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            table_data = []
            for tbl in tables:
                ths = tbl.find_elements(By.TAG_NAME, "th")
                trs = tbl.find_elements(By.TAG_NAME, "tr")
                headers = [th.text.strip() for th in ths]
                table_data.append({
                    "headers": headers,
                    "rows": len(trs) - 1
                })

            self.log("INFO", f"Tabelas: {len(table_data)}")

            self.data["secoes"]["Qualidade"] = {
                "url": url_atual,
                "titulo": titulo,
                "formularios": form_count,
                "inputs": input_data[:20],  # Primeiros 20
                "selects": select_data[:20],
                "tabelas": table_data[:10]
            }

            self.screenshot("secao_qualidade.png")

        except Exception as e:
            self.log("AVISO", f"Extracao: {str(e)[:80]}")

    def screenshot(self, nome):
        try:
            os.makedirs("screenshots", exist_ok=True)
            caminho = os.path.join("screenshots", nome)
            self.driver.save_screenshot(caminho)
            self.log("INFO", f"Screenshot: {nome}")
        except:
            pass

    def exportar_json(self):
        try:
            with open("relatorio_mobuss_qualidade.json", 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            self.log("OK", "JSON exportado")
        except Exception as e:
            self.log("ERRO", f"JSON: {e}")

    def exportar_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Qualidade"

            row = 1
            for secao, dados in self.data["secoes"].items():
                ws[f'A{row}'] = "Secao"
                ws[f'B{row}'] = secao
                row += 1

                ws[f'A{row}'] = "URL"
                ws[f'B{row}'] = dados.get("url", "")
                row += 1

                ws[f'A{row}'] = "Titulo"
                ws[f'B{row}'] = dados.get("titulo", "")
                row += 1

                ws[f'A{row}'] = "Formularios"
                ws[f'B{row}'] = dados.get("formularios", 0)
                row += 1

                ws[f'A{row}'] = "Inputs"
                ws[f'B{row}'] = len(dados.get("inputs", []))
                row += 1

                ws[f'A{row}'] = "Selects"
                ws[f'B{row}'] = len(dados.get("selects", []))
                row += 1

                ws[f'A{row}'] = "Tabelas"
                ws[f'B{row}'] = len(dados.get("tabelas", []))
                row += 2

                # Inputs
                if dados.get("inputs"):
                    ws[f'A{row}'] = "=== INPUTS ==="
                    row += 1
                    for inp in dados.get("inputs", []):
                        ws[f'A{row}'] = inp.get("type", "")
                        ws[f'B{row}'] = inp.get("name", "")
                        ws[f'C{row}'] = inp.get("id", "")
                        row += 1

                row += 1

                # Selects
                if dados.get("selects"):
                    ws[f'A{row}'] = "=== SELECTS ==="
                    row += 1
                    for sel in dados.get("selects", []):
                        ws[f'A{row}'] = sel.get("name", "")
                        ws[f'B{row}'] = ", ".join(sel.get("options", [])[:5])
                        row += 1

            wb.save("relatorio_mobuss_qualidade.xlsx")
            self.log("OK", "Excel exportado")

        except Exception as e:
            self.log("ERRO", f"Excel: {e}")

    def executar(self):
        try:
            self.log("INFO", "=== SCRAPER MOBUSS QUALIDADE ===\n")

            self.setup_driver()
            self.login()
            time.sleep(2)
            self.acessar_qualidade()
            time.sleep(2)
            self.extrair_dados()

            self.log("INFO", "\n=== EXPORTANDO ===\n")
            self.exportar_json()
            self.exportar_excel()

            self.log("OK", "\nConcluido com sucesso!")

        except Exception as e:
            self.log("ERRO", f"Geral: {e}")
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass

if __name__ == "__main__":
    scraper = MobussScraper()
    scraper.executar()
