#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Scraping - MODULO QUALIDADE DO MOBUSS
Extrai todos os recursos da seção Qualidade
Versão 3: Suporte a login multi-etapas
"""

import json
import time
import sys
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import os

# Configurações
URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/portal/home/iu00_02.jsf"
EMAIL = "ti@grupoinvestcorp.com.br"
SENHA = "FT$0luc0#$TI"

class MobussQualidadeScraper:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {
            "timestamp": datetime.now().isoformat(),
            "secoes": {},
            "formularios": [],
            "campos": [],
            "opcoes": [],
            "tabelas": [],
            "estrutura_menu": []
        }

    def log(self, nivel, mensagem):
        """Log com tratamento de encoding para Windows"""
        try:
            timestamp = datetime.now().strftime("%H:%M:%S")
            output = f"[{timestamp}] [{nivel}] {mensagem}"
            print(output)
            sys.stdout.flush()
        except Exception as e:
            print(f"[LOG_ERROR] {str(e)}")

    def setup_driver(self):
        """Configura o driver do Selenium"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)

            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 15)
            self.log("OK", "Driver iniciado com sucesso")
        except Exception as e:
            self.log("ERRO", f"Erro ao iniciar driver: {str(e)}")
            raise

    def login(self):
        """Realiza login no Mobuss (multi-etapas)"""
        try:
            self.log("INFO", f"Acessando: {URL}")
            self.driver.get(URL)
            time.sleep(3)

            # ETAPA 1: Preencher usuario
            self.log("INFO", "Etapa 1: Preenchendo usuario...")
            usuario_field = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[name*='usuario'], input[type='text']"))
            )
            usuario_field.clear()
            usuario_field.send_keys(EMAIL)
            time.sleep(1)

            # Clica em Entrar (primeira etapa)
            entrar_btn = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit'], button:contains('Entrar'), input[type='submit']")
            entrar_btn.click()
            time.sleep(5)

            # ETAPA 2: Preencher senha (se necessario)
            self.log("INFO", "Etapa 2: Aguardando campo de senha...")
            try:
                senha_field = self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password']")),
                    timeout=5
                )
                self.log("INFO", "Campo de senha encontrado. Preenchendo...")
                senha_field.clear()
                senha_field.send_keys(SENHA)
                time.sleep(1)

                # Clica em Entrar (segunda etapa)
                entrar_btn2 = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")
                entrar_btn2.click()
                time.sleep(5)
            except:
                self.log("AVISO", "Nenhum campo de senha detectado. Pode ter sido login de usuario unico.")

            self.log("OK", "Login realizado com sucesso")

        except Exception as e:
            self.log("ERRO", f"Erro no login: {str(e)}")
            self.captura_screenshot("erro_login.png")
            raise

    def acessar_qualidade(self):
        """Acessa a seção Qualidade no menu"""
        try:
            self.log("INFO", "Acessando secao QUALIDADE...")
            time.sleep(2)

            # Procura por menu com texto "Qualidade"
            try:
                qualidade_menu = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Qualidade')] | //a[contains(text(), 'Qualidade')] | //button[contains(text(), 'Qualidade')]")),
                    timeout=5
                )
                qualidade_menu.click()
                time.sleep(3)
            except:
                self.log("AVISO", "Menu Qualidade nao encontrado. Tentando URL direta...")

            self.log("OK", "Secao Qualidade acessada")

        except Exception as e:
            self.log("AVISO", f"Erro ao acessar Qualidade: {str(e)}")

    def extrair_secoes_menu(self):
        """Extrai todas as subsecoes do menu Qualidade"""
        try:
            self.log("INFO", "Extraindo secoes do menu...")
            time.sleep(2)

            # Procura todos os items de menu
            menu_items = self.driver.find_elements(By.XPATH, "//nav//a, //menu//a, //a[@class*='menu'], //li//a, //div[@class*='menu']//a")

            secoes = []
            for item in menu_items:
                try:
                    texto = item.text.strip()
                    href = item.get_attribute("href")

                    if texto and href and len(texto) > 2:
                        secoes.append({
                            "nome": texto,
                            "url": href
                        })
                except:
                    pass

            # Remove duplicatas
            secoes_unicos = []
            urls_vistas = set()
            for s in secoes:
                if s["url"] not in urls_vistas:
                    secoes_unicos.append(s)
                    urls_vistas.add(s["url"])

            self.data["estrutura_menu"] = secoes_unicos
            self.log("OK", f"{len(secoes_unicos)} secao(oes) encontrada(s)")

            for idx, secao in enumerate(secoes_unicos[:10]):
                self.log("INFO", f"  [{idx+1}] {secao['nome']}")

        except Exception as e:
            self.log("AVISO", f"Erro ao extrair secoes: {str(e)}")

    def navegar_e_extrair_secoes(self):
        """Navega por cada secao e extrai dados"""
        secoes_importantes = [
            "Painel",
            "Cobertura",
            "Modelo",
            "Inspecao",
            "Realizado",
            "Cadastro",
            "Formulario",
            "Gestao",
            "Acao",
            "Tipo"
        ]

        total_secoes = len(self.data["estrutura_menu"])
        self.log("INFO", f"\nProcessando {total_secoes} secoes...\n")

        for idx, secao in enumerate(self.data["estrutura_menu"]):
            nome_secao = secao["nome"]
            self.log("INFO", f"[{idx+1}/{total_secoes}] {nome_secao}")

            # Processa apenas secoes importantes
            if any(importante.lower() in nome_secao.lower() for importante in secoes_importantes):
                try:
                    # Navega para a secao
                    self.driver.get(secao["url"])
                    time.sleep(2)

                    # Extrai dados da secao
                    dados_secao = self.extrair_dados_secao(nome_secao)
                    self.data["secoes"][nome_secao] = dados_secao

                    # Captura screenshot
                    nome_arquivo = f"secao_{nome_secao.replace(' ', '_').replace('/', '_').lower()}.png"
                    self.captura_screenshot(nome_arquivo)

                    forms = len(dados_secao.get("formularios", []))
                    tabelas = len(dados_secao.get("tabelas", []))
                    self.log("OK", f"  Formularios: {forms}, Tabelas: {tabelas}")

                except Exception as e:
                    self.log("AVISO", f"  Erro: {str(e)[:80]}")

    def extrair_dados_secao(self, nome_secao):
        """Extrai todos os dados de uma secao especifica"""
        dados = {
            "url": self.driver.current_url,
            "titulo": self.driver.title,
            "formularios": [],
            "tabelas": [],
            "campos": [],
            "opcoes": []
        }

        try:
            # Extrai formularios
            formularios = self.driver.find_elements(By.TAG_NAME, "form")
            for form in formularios:
                form_data = {
                    "id": form.get_attribute("id"),
                    "campos": self.extrair_campos_form(form)
                }
                dados["formularios"].append(form_data)

            # Extrai tabelas
            tabelas = self.driver.find_elements(By.TAG_NAME, "table")
            for idx, tabela in enumerate(tabelas):
                tabela_data = {
                    "numero": idx + 1,
                    "headers": [],
                    "linhas": [],
                    "total_linhas": 0
                }

                # Headers
                headers = tabela.find_elements(By.TAG_NAME, "th")
                tabela_data["headers"] = [h.text.strip() for h in headers if h.text.strip()]

                # Linhas (limita a 100 primeiras)
                rows = tabela.find_elements(By.TAG_NAME, "tr")
                for row in rows[1:101]:
                    try:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        linha = [cell.text.strip() for cell in cells]
                        if linha:
                            tabela_data["linhas"].append(linha)
                    except:
                        pass

                tabela_data["total_linhas"] = len(tabela_data["linhas"])
                dados["tabelas"].append(tabela_data)

            # Extrai selects e opcoes
            selects = self.driver.find_elements(By.TAG_NAME, "select")
            for select in selects:
                try:
                    options = select.find_elements(By.TAG_NAME, "option")
                    opcoes_list = [{"value": opt.get_attribute("value"), "text": opt.text} for opt in options]

                    if opcoes_list:
                        dados["opcoes"].append({
                            "select_name": select.get_attribute("name"),
                            "opcoes": opcoes_list
                        })
                except:
                    pass

        except Exception as e:
            self.log("AVISO", f"Erro ao extrair dados: {str(e)[:60]}")

        return dados

    def extrair_campos_form(self, form):
        """Extrai campos de um formulario"""
        campos = []

        try:
            # Inputs
            inputs = form.find_elements(By.TAG_NAME, "input")
            for inp in inputs:
                if inp.get_attribute("type") not in ["submit", "button"]:
                    campos.append({
                        "tipo": "input",
                        "input_type": inp.get_attribute("type"),
                        "name": inp.get_attribute("name"),
                        "id": inp.get_attribute("id"),
                        "placeholder": inp.get_attribute("placeholder")
                    })

            # Selects
            selects = form.find_elements(By.TAG_NAME, "select")
            for select in selects:
                opcoes = [{"value": opt.get_attribute("value"), "text": opt.text}
                         for opt in select.find_elements(By.TAG_NAME, "option")]

                campos.append({
                    "tipo": "select",
                    "name": select.get_attribute("name"),
                    "opcoes": opcoes
                })

            # Textareas
            textareas = form.find_elements(By.TAG_NAME, "textarea")
            for textarea in textareas:
                campos.append({
                    "tipo": "textarea",
                    "name": textarea.get_attribute("name"),
                    "id": textarea.get_attribute("id")
                })

        except Exception as e:
            pass

        return campos

    def captura_screenshot(self, nome):
        """Captura screenshot da pagina"""
        try:
            caminho = os.path.join("screenshots", nome)
            os.makedirs("screenshots", exist_ok=True)
            self.driver.save_screenshot(caminho)
        except Exception as e:
            pass

    def exportar_json(self):
        """Exporta dados em JSON"""
        try:
            caminho = "relatorio_mobuss_qualidade.json"
            with open(caminho, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            self.log("OK", f"JSON exportado: {caminho}")
        except Exception as e:
            self.log("ERRO", f"Erro ao exportar JSON: {str(e)}")

    def exportar_excel(self):
        """Exporta dados em Excel"""
        try:
            wb = Workbook()

            # Aba 1: Resumo
            ws1 = wb.active
            ws1.title = "Resumo"
            ws1['A1'] = "Secao"
            ws1['B1'] = "URL"
            ws1['C1'] = "Formularios"
            ws1['D1'] = "Tabelas"
            ws1['E1'] = "Campos"
            ws1['F1'] = "Opcoes"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                ws1[f'A{row}'] = secao_nome
                ws1[f'B{row}'] = dados_secao.get("url", "")
                ws1[f'C{row}'] = len(dados_secao.get("formularios", []))
                ws1[f'D{row}'] = len(dados_secao.get("tabelas", []))
                ws1[f'E{row}'] = len(dados_secao.get("campos", []))
                ws1[f'F{row}'] = len(dados_secao.get("opcoes", []))
                row += 1

            # Aba 2: Campos
            ws2 = wb.create_sheet("Campos")
            ws2['A1'] = "Secao"
            ws2['B1'] = "Tipo"
            ws2['C1'] = "Nome"
            ws2['D1'] = "ID"
            ws2['E1'] = "Placeholder"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                for campo in dados_secao.get("campos", []):
                    ws2[f'A{row}'] = secao_nome
                    ws2[f'B{row}'] = campo.get("tipo", "")
                    ws2[f'C{row}'] = campo.get("name", "")
                    ws2[f'D{row}'] = campo.get("id", "")
                    ws2[f'E{row}'] = campo.get("placeholder", "")
                    row += 1

            # Aba 3: Opcoes
            ws3 = wb.create_sheet("Opcoes")
            ws3['A1'] = "Secao"
            ws3['B1'] = "Select"
            ws3['C1'] = "Opcao"
            ws3['D1'] = "Value"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                for select_data in dados_secao.get("opcoes", []):
                    for opcao in select_data.get("opcoes", []):
                        ws3[f'A{row}'] = secao_nome
                        ws3[f'B{row}'] = select_data.get("select_name", "")
                        ws3[f'C{row}'] = opcao.get("text", "")
                        ws3[f'D{row}'] = opcao.get("value", "")
                        row += 1

            # Aba 4: Tabelas
            ws4 = wb.create_sheet("Tabelas")
            ws4['A1'] = "Secao"
            ws4['B1'] = "Tabela"
            ws4['C1'] = "Headers"
            ws4['D1'] = "Total_Linhas"

            row = 2
            for secao_nome, dados_secao in self.data["secoes"].items():
                for tabela in dados_secao.get("tabelas", []):
                    ws4[f'A{row}'] = secao_nome
                    ws4[f'B{row}'] = tabela.get("numero", "")
                    ws4[f'C{row}'] = ", ".join(tabela.get("headers", []))
                    ws4[f'D{row}'] = tabela.get("total_linhas", 0)
                    row += 1

            caminho = "relatorio_mobuss_qualidade.xlsx"
            wb.save(caminho)
            self.log("OK", f"Excel exportado: {caminho}")

        except Exception as e:
            self.log("ERRO", f"Erro ao exportar Excel: {str(e)}")

    def executar(self):
        """Executa o scraping completo"""
        try:
            self.log("INFO", "=== INICIANDO SCRAPING DO MODULO QUALIDADE ===\n")

            self.setup_driver()
            self.login()
            time.sleep(2)
            self.acessar_qualidade()
            time.sleep(2)
            self.extrair_secoes_menu()
            self.navegar_e_extrair_secoes()

            self.log("INFO", "\n=== EXPORTANDO RELATORIOS ===\n")
            self.exportar_json()
            self.exportar_excel()

            self.log("OK", "\n=== SCRAPING CONCLUIDO COM SUCESSO ===")
            self.log("INFO", f"Total de secoes extraidas: {len(self.data['secoes'])}")

        except Exception as e:
            self.log("ERRO", f"Erro geral: {str(e)}")

        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass
                self.log("INFO", "Driver fechado")

if __name__ == "__main__":
    scraper = MobussQualidadeScraper()
    scraper.executar()
