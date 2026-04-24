#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Scraping - MODULO QUALIDADE DO MOBUSS
Versao final com seletores corrigidos
"""

import json
import time
import sys
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import os

URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/portal/home/iu00_02.jsf"
EMAIL = "ti@grupoinvestcorp.com.br"
SENHA = "FT$0luc0#$TI"

class MobussQualidadeScraper:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {
            "timestamp": datetime.now().isoformat(),
            "secoes": {},
            "estrutura_menu": []
        }

    def log(self, nivel, mensagem):
        timestamp = datetime.now().strftime("%H:%M:%S")
        output = f"[{timestamp}] [{nivel}] {mensagem}"
        print(output)
        sys.stdout.flush()

    def setup_driver(self):
        try:
            chrome_options = Options()
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")

            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 15)
            self.log("OK", "Driver iniciado")
        except Exception as e:
            self.log("ERRO", f"Setup driver: {str(e)}")
            raise

    def login(self):
        try:
            self.log("INFO", f"Acessando: {URL}")
            self.driver.get(URL)
            time.sleep(3)

            # Procura campo usuario
            self.log("INFO", "Buscando campo de usuario...")
            input_fields = self.driver.find_elements(By.TAG_NAME, "input")

            usuario_field = None
            for inp in input_fields:
                field_type = inp.get_attribute("type")
                field_name = inp.get_attribute("name")
                if field_type == "text" or "usuario" in (field_name or "").lower():
                    usuario_field = inp
                    break

            if not usuario_field:
                usuario_field = input_fields[0]

            self.log("INFO", "Preenchendo usuario...")
            usuario_field.clear()
            usuario_field.send_keys(EMAIL)
            time.sleep(1)

            # Clica botao
            buttons = self.driver.find_elements(By.TAG_NAME, "button")
            if buttons:
                buttons[0].click()
            else:
                inputs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='submit']")
                if inputs:
                    inputs[0].click()

            time.sleep(5)
            self.log("OK", "Login realizado")

        except Exception as e:
            self.log("ERRO", f"Login: {str(e)[:100]}")
            self.captura_screenshot("erro_login.png")
            raise

    def acessar_qualidade(self):
        try:
            self.log("INFO", "Acessando Qualidade...")
            time.sleep(2)

            links = self.driver.find_elements(By.TAG_NAME, "a")
            for link in links:
                if "Qualidade" in (link.text or ""):
                    link.click()
                    time.sleep(3)
                    break

            self.log("OK", "Qualidade acessada")
        except Exception as e:
            self.log("AVISO", f"Qualidade: {str(e)[:80]}")

    def extrair_secoes_menu(self):
        try:
            self.log("INFO", "Extraindo menu...")

            links = self.driver.find_elements(By.TAG_NAME, "a")
            secoes = []

            for link in links:
                texto = (link.text or "").strip()
                href = link.get_attribute("href")

                if texto and href and len(texto) > 2:
                    secoes.append({"nome": texto, "url": href})

            # Remove duplicatas
            seen = set()
            unique = []
            for s in secoes:
                if s["url"] not in seen:
                    seen.add(s["url"])
                    unique.append(s)

            self.data["estrutura_menu"] = unique
            self.log("OK", f"{len(unique)} secoes encontradas")

            for idx, s in enumerate(unique[:15]):
                self.log("INFO", f"  [{idx+1}] {s['nome'][:50]}")

        except Exception as e:
            self.log("AVISO", f"Menu: {str(e)[:80]}")

    def extrair_secao(self, url, nome):
        try:
            self.driver.get(url)
            time.sleep(2)

            # Extrair formularios
            forms = self.driver.find_elements(By.TAG_NAME, "form")
            form_count = len(forms)

            # Extrair tabelas
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            table_data = []
            for table in tables:
                headers = [th.text.strip() for th in table.find_elements(By.TAG_NAME, "th")]
                rows = len(table.find_elements(By.TAG_NAME, "tr"))
                table_data.append({"headers": headers, "rows": rows})

            # Extrair campos
            inputs = self.driver.find_elements(By.TAG_NAME, "input")
            selects = self.driver.find_elements(By.TAG_NAME, "select")

            campo_count = len(inputs) + len(selects)

            self.data["secoes"][nome] = {
                "url": url,
                "formularios": form_count,
                "tabelas": len(table_data),
                "campos": campo_count,
                "tabelas_detalhes": table_data
            }

            self.captura_screenshot(f"secao_{nome.replace(' ', '_')}.png")
            self.log("OK", f"{nome}: Forms={form_count}, Tables={len(table_data)}, Fields={campo_count}")

        except Exception as e:
            self.log("AVISO", f"{nome}: {str(e)[:60]}")

    def navegar_secoes(self):
        self.log("INFO", "\nProcessando secoes...")

        for idx, secao in enumerate(self.data["estrutura_menu"][:12]):
            nome = secao["nome"]
            url = secao["url"]

            self.log("INFO", f"[{idx+1}] {nome[:40]}...")
            self.extrair_secao(url, nome)

    def captura_screenshot(self, nome):
        try:
            caminho = os.path.join("screenshots", nome)
            os.makedirs("screenshots", exist_ok=True)
            self.driver.save_screenshot(caminho)
        except:
            pass

    def exportar_json(self):
        try:
            with open("relatorio_mobuss_qualidade.json", 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            self.log("OK", "JSON exportado")
        except Exception as e:
            self.log("ERRO", f"JSON: {str(e)}")

    def exportar_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Qualidade"

            ws['A1'] = "Secao"
            ws['B1'] = "URL"
            ws['C1'] = "Formularios"
            ws['D1'] = "Tabelas"
            ws['E1'] = "Campos"

            row = 2
            for nome, dados in self.data["secoes"].items():
                ws[f'A{row}'] = nome
                ws[f'B{row}'] = dados.get("url", "")
                ws[f'C{row}'] = dados.get("formularios", 0)
                ws[f'D{row}'] = dados.get("tabelas", 0)
                ws[f'E{row}'] = dados.get("campos", 0)
                row += 1

            wb.save("relatorio_mobuss_qualidade.xlsx")
            self.log("OK", "Excel exportado")

        except Exception as e:
            self.log("ERRO", f"Excel: {str(e)}")

    def executar(self):
        try:
            self.log("INFO", "=== SCRAPING MOBUSS QUALIDADE ===\n")

            self.setup_driver()
            self.login()
            time.sleep(2)
            self.acessar_qualidade()
            time.sleep(2)
            self.extrair_secoes_menu()
            self.navegar_secoes()

            self.log("INFO", "\n=== EXPORTANDO ===\n")
            self.exportar_json()
            self.exportar_excel()

            self.log("OK", f"\nConcluido! {len(self.data['secoes'])} secoes extraidas")

        except Exception as e:
            self.log("ERRO", f"Geral: {str(e)}")
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass
                self.log("INFO", "Driver fechado")

if __name__ == "__main__":
    scraper = MobussQualidadeScraper()
    scraper.executar()
