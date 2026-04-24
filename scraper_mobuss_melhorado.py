#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scraper Mobuss - Fluxo de login em duas etapas
"""

import json
import time
import sys
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import os

LOGIN_URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/portal/home/iu00_02.jsf"
QUALIDADE_URL = "https://www.mobuss.com.br/ccweb/htdocs/programs/qualidade/checklist/realizado/iu01_23.jsf"

EMAIL = "ti@grupoinvestcorp.com.br"
SENHA = "FT$0luc0#$TI"

class MobussScraper:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {
            "timestamp": datetime.now().isoformat(),
            "secoes": {}
        }

    def log(self, nivel, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] [{nivel}] {msg}")
        sys.stdout.flush()

    def setup_driver(self):
        try:
            opts = Options()
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            self.driver = webdriver.Chrome(options=opts)
            self.wait = WebDriverWait(self.driver, 15)
            self.log("OK", "Driver iniciado")
        except Exception as e:
            self.log("ERRO", f"Setup: {e}")
            raise

    def login(self):
        """Login em duas etapas: usuario -> senha"""
        try:
            self.log("INFO", f"Etapa 1: Acessando login")
            self.driver.get(LOGIN_URL)
            time.sleep(3)

            # ETAPA 1: Preencher username
            self.log("INFO", "Etapa 1: Buscando campo de usuario...")
            inputs = self.driver.find_elements(By.TAG_NAME, "input")

            user_input = None
            for inp in inputs:
                inp_type = inp.get_attribute("type")
                inp_name = inp.get_attribute("name")

                if inp_type == "text" and ("user" in (inp_name or "").lower() or "email" in (inp_name or "").lower()):
                    user_input = inp
                    break

            if not user_input:
                # Tenta o primeiro input type=text
                for inp in inputs:
                    if inp.get_attribute("type") == "text":
                        user_input = inp
                        break

            if user_input:
                self.log("INFO", f"Campo usuario encontrado. Preenchendo: {EMAIL}")
                user_input.clear()
                user_input.send_keys(EMAIL)
                time.sleep(1)

                # Clica botao Entrar
                self.log("INFO", "Etapa 1: Clicando em Entrar...")
                btns = self.driver.find_elements(By.TAG_NAME, "button")
                if btns:
                    btns[0].click()
                else:
                    subs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='submit']")
                    if subs:
                        subs[0].click()

                time.sleep(4)
                self.log("OK", "Etapa 1 concluida")
            else:
                self.log("ERRO", "Campo usuario nao encontrado")
                self.screenshot("erro_campo_usuario.png")
                raise Exception("Campo usuario nao encontrado")

            # ETAPA 2: Preencher password
            self.log("INFO", "Etapa 2: Buscando campo de senha...")
            time.sleep(2)

            pwd_inputs = self.driver.find_elements(By.TAG_NAME, "input")
            pwd_input = None

            for inp in pwd_inputs:
                inp_type = inp.get_attribute("type")
                if inp_type == "password":
                    pwd_input = inp
                    break

            if pwd_input:
                self.log("INFO", "Campo senha encontrado. Preenchendo...")
                pwd_input.clear()
                pwd_input.send_keys(SENHA)
                time.sleep(1)

                # Clica botao Entrar (segunda vez)
                self.log("INFO", "Etapa 2: Clicando em Entrar...")
                btns = self.driver.find_elements(By.TAG_NAME, "button")
                if btns:
                    btns[0].click()
                else:
                    subs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='submit']")
                    if subs:
                        subs[0].click()

                time.sleep(5)
                self.log("OK", "Etapa 2 concluida - autenticacao completa")

                # Verifica se ainda esta na pagina de login
                if "Entrar" in self.driver.title or "login" in self.driver.current_url.lower():
                    self.log("AVISO", "Ainda na pagina de login. Tentando novamente...")
                    self.screenshot("aviso_login_falhou.png")
                    time.sleep(2)
                else:
                    self.log("OK", "Autenticacao bem-sucedida!")
            else:
                self.log("AVISO", "Campo senha nao encontrado. Pode ja estar autenticado.")
                self.screenshot("aviso_sem_senha.png")

        except Exception as e:
            self.log("ERRO", f"Login: {str(e)[:150]}")
            self.screenshot("erro_login_geral.png")
            raise

    def acessar_qualidade(self):
        try:
            self.log("INFO", "Acessando Qualidade...")

            # Acessa URL direta se nao estiver la
            url_atual = self.driver.current_url.lower()
            if "qualidade" not in url_atual:
                self.log("INFO", f"Navegando para Qualidade: {QUALIDADE_URL}")
                self.driver.get(QUALIDADE_URL)

            time.sleep(4)

            # Verifica se conseguiu acessar
            titulo = self.driver.title
            url = self.driver.current_url

            self.log("INFO", f"URL atual: {url[:80]}")
            self.log("INFO", f"Titulo: {titulo[:80]}")

            if "Entrar" in titulo or "login" in url.lower():
                self.log("AVISO", "Redirecionado para login. Sessao pode ter expirado.")
            else:
                self.log("OK", "Qualidade acessada com sucesso")

        except Exception as e:
            self.log("ERRO", f"Qualidade: {e}")
            raise

    def extrair_dados(self):
        try:
            self.log("INFO", "Extraindo dados de Qualidade...")

            titulo = self.driver.title
            url = self.driver.current_url

            # Formularios
            forms = self.driver.find_elements(By.TAG_NAME, "form")
            form_count = len(forms)
            self.log("INFO", f"Formularios encontrados: {form_count}")

            # Inputs
            all_inputs = self.driver.find_elements(By.TAG_NAME, "input")
            input_data = []
            for inp in all_inputs:
                itype = inp.get_attribute("type")
                name = inp.get_attribute("name")
                placeholder = inp.get_attribute("placeholder")

                if itype not in ["submit", "button", "hidden"] and name:
                    input_data.append({
                        "type": itype,
                        "name": name,
                        "id": inp.get_attribute("id"),
                        "placeholder": placeholder
                    })

            self.log("INFO", f"Campos input: {len(input_data)}")

            # Selects
            selects = self.driver.find_elements(By.TAG_NAME, "select")
            select_data = []
            for sel in selects:
                opts = sel.find_elements(By.TAG_NAME, "option")
                opt_texts = [opt.text.strip() for opt in opts if opt.text.strip()]
                select_data.append({
                    "name": sel.get_attribute("name"),
                    "id": sel.get_attribute("id"),
                    "options": opt_texts
                })

            self.log("INFO", f"Selects: {len(select_data)}")

            # Tabelas
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            table_data = []
            for tbl in tables:
                ths = tbl.find_elements(By.TAG_NAME, "th")
                trs = tbl.find_elements(By.TAG_NAME, "tr")
                headers = [th.text.strip() for th in ths if th.text.strip()]
                table_data.append({
                    "headers": headers,
                    "rows": len(trs) - 1 if trs else 0
                })

            self.log("INFO", f"Tabelas: {len(table_data)}")

            # Textareas
            textareas = self.driver.find_elements(By.TAG_NAME, "textarea")
            self.log("INFO", f"Textareas: {len(textareas)}")

            self.data["secoes"]["Qualidade - Realizados"] = {
                "url": url,
                "titulo": titulo,
                "formularios": form_count,
                "inputs": input_data[:30],
                "selects": select_data[:20],
                "tabelas": table_data[:10],
                "textareas": len(textareas)
            }

            self.screenshot("secao_qualidade_dados.png")
            self.log("OK", "Dados extraidos com sucesso")

        except Exception as e:
            self.log("AVISO", f"Extracao: {str(e)[:100]}")

    def screenshot(self, nome):
        try:
            os.makedirs("screenshots", exist_ok=True)
            caminho = os.path.join("screenshots", nome)
            self.driver.save_screenshot(caminho)
            self.log("INFO", f"Screenshot: {nome}")
        except:
            pass

    def exportar_json(self):
        try:
            with open("relatorio_mobuss_qualidade.json", 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            self.log("OK", "JSON exportado")
        except Exception as e:
            self.log("ERRO", f"JSON: {e}")

    def exportar_excel(self):
        try:
            wb = Workbook()

            # Aba 1: Resumo
            ws1 = wb.active
            ws1.title = "Resumo"
            ws1['A1'] = "Parametro"
            ws1['B1'] = "Valor"

            row = 2
            for secao, dados in self.data["secoes"].items():
                ws1[f'A{row}'] = "Secao"
                ws1[f'B{row}'] = secao
                row += 1

                ws1[f'A{row}'] = "URL"
                ws1[f'B{row}'] = dados.get("url", "")
                row += 1

                ws1[f'A{row}'] = "Formularios"
                ws1[f'B{row}'] = dados.get("formularios", 0)
                row += 1

                ws1[f'A{row}'] = "Inputs"
                ws1[f'B{row}'] = len(dados.get("inputs", []))
                row += 1

                ws1[f'A{row}'] = "Selects"
                ws1[f'B{row}'] = len(dados.get("selects", []))
                row += 1

                ws1[f'A{row}'] = "Tabelas"
                ws1[f'B{row}'] = len(dados.get("tabelas", []))
                row += 1

                ws1[f'A{row}'] = "Textareas"
                ws1[f'B{row}'] = dados.get("textareas", 0)
                row += 2

            # Aba 2: Campos Input
            ws2 = wb.create_sheet("Inputs")
            ws2['A1'] = "Secao"
            ws2['B1'] = "Tipo"
            ws2['C1'] = "Nome"
            ws2['D1'] = "ID"
            ws2['E1'] = "Placeholder"

            row = 2
            for secao, dados in self.data["secoes"].items():
                for inp in dados.get("inputs", []):
                    ws2[f'A{row}'] = secao
                    ws2[f'B{row}'] = inp.get("type", "")
                    ws2[f'C{row}'] = inp.get("name", "")
                    ws2[f'D{row}'] = inp.get("id", "")
                    ws2[f'E{row}'] = inp.get("placeholder", "")
                    row += 1

            # Aba 3: Selects
            ws3 = wb.create_sheet("Selects")
            ws3['A1'] = "Secao"
            ws3['B1'] = "Nome"
            ws3['C1'] = "ID"
            ws3['D1'] = "Opcoes"

            row = 2
            for secao, dados in self.data["secoes"].items():
                for sel in dados.get("selects", []):
                    ws3[f'A{row}'] = secao
                    ws3[f'B{row}'] = sel.get("name", "")
                    ws3[f'C{row}'] = sel.get("id", "")
                    ws3[f'D{row}'] = "; ".join(sel.get("options", [])[:10])
                    row += 1

            # Aba 4: Tabelas
            ws4 = wb.create_sheet("Tabelas")
            ws4['A1'] = "Secao"
            ws4['B1'] = "Headers"
            ws4['C1'] = "Linhas"

            row = 2
            for secao, dados in self.data["secoes"].items():
                for idx, tbl in enumerate(dados.get("tabelas", []), 1):
                    ws4[f'A{row}'] = f"{secao} - Tabela {idx}"
                    ws4[f'B{row}'] = "; ".join(tbl.get("headers", []))
                    ws4[f'C{row}'] = tbl.get("rows", 0)
                    row += 1

            wb.save("relatorio_mobuss_qualidade.xlsx")
            self.log("OK", "Excel exportado")

        except Exception as e:
            self.log("ERRO", f"Excel: {e}")

    def executar(self):
        try:
            self.log("INFO", "=== SCRAPER MOBUSS QUALIDADE v2.0 ===\n")

            self.setup_driver()
            self.login()
            time.sleep(2)
            self.acessar_qualidade()
            time.sleep(2)
            self.extrair_dados()

            self.log("INFO", "\n=== EXPORTANDO RELATORIOS ===\n")
            self.exportar_json()
            self.exportar_excel()

            self.log("OK", "\nScraping concluido com sucesso!")

        except Exception as e:
            self.log("ERRO", f"Execucao geral: {e}")
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass

if __name__ == "__main__":
    scraper = MobussScraper()
    scraper.executar()
